import openpyxl
import os
def ölç():
    a=1
    for a in range(1,1000):
        b=1
        for b in range(1,1000):
            if((sheet.cell(row=b,column=a).value==None)):
                break
            print(sheet.cell(row=b,column=a).value)
        b=1
        if(sheet.cell(row=b,column=a+1).value==None):
            break
s=""
h=["A","B","C","D"]
def sifirla(row,col):
    a=1
    for a in range(1,row):
        b=1
        for b in range(1,col):
            s=h[b-1]+str(b)
            print(sheet[s].value)
            sheet[s]=""
            s=""
os.chdir('C:/Users/pc/Desktop')
workbook = openpyxl.load_workbook('a.xlsx')
sheet = workbook['Sayfa1']
#ölç()
sifirla(10,4)
#ölç()
workbook.save('b.xlsx')
